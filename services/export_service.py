"""
Service: Xuất dữ liệu Excel và PDF.
Dùng openpyxl cho .xlsx, reportlab cho .pdf
"""

import io
from datetime import datetime

from db.connection import execute_query
from services import class_service


def _create_attendance_pie_chart(present, late, absent):
    import matplotlib

    matplotlib.use("Agg")  # Tránh lỗi main thread GUI trên server
    import matplotlib.pyplot as plt
    import io

    labels = []
    sizes = []
    colors = []

    if present > 0:
        labels.append("Co mat")
        sizes.append(present)
        colors.append("#2ecc71")  # Green
    if late > 0:
        labels.append("Di tre")
        sizes.append(late)
        colors.append("#f39c12")  # Orange
    if absent > 0:
        labels.append("Vang mat")
        sizes.append(absent)
        colors.append("#e74c3c")  # Red

    if not sizes:
        labels = ["Chua co du lieu"]
        sizes = [1]
        colors = ["#bdc3c7"]

    fig, ax = plt.subplots(figsize=(6, 3.5))  # Tăng kích thước chút cho đẹp
    patches = ax.pie(
        sizes,
        colors=colors,
        labels=labels,
        autopct="%1.1f%%",
        pctdistance=0.75,
        startangle=90,
        textprops=dict(color="#333333", weight="bold", fontsize=10),
        wedgeprops=dict(width=0.4, edgecolor="white", linewidth=2),  # Donut chart
    )

    ax.axis("equal")

    buf = io.BytesIO()
    plt.savefig(buf, format="png", transparent=True, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def to_excel(lop_id, date=None):
    """
    Tạo file Excel (.xlsx) điểm danh của lớp.
    Query theo phiên (phien_id) thay vì theo ngày.

    Args:
        lop_id: ID lớp học
        date: Ngày cần xuất (YYYY-MM-DD), mặc định hôm nay

    Returns:
        bytes: Nội dung file Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as ExcelImage
    from services.attendance_policy import AttendanceStatus

    lop = class_service.get_by_id(lop_id)
    if not lop:
        return None

    if not date:
        date = datetime.now().strftime("%Y-%m-%d")

    # Tìm phiên đã chốt trong ngày cho lớp này
    session_row = execute_query(
        """SELECT id FROM phien_diem_danh
           WHERE lop_id = %s AND CAST(bat_dau AS DATE) = %s
           AND ISNULL(is_cancelled, 0) = 0
           ORDER BY bat_dau DESC""",
        (lop_id, date),
    )

    if session_row:
        phien_ids = [s["id"] for s in session_row]
        placeholders = ",".join(["%s"] * len(phien_ids))
        # Query dữ liệu điểm danh — lấy từ DB, dùng status enum
        sql = f"""
            SELECT sv.mssv, sv.ho_ten,
                   dd.thoi_gian, dd.status, dd.late_minutes, dd.gio_vao_lop,
                   dd.do_chinh_xac
            FROM sinh_vien sv
            LEFT JOIN diem_danh dd ON sv.id = dd.sinh_vien_id
                AND dd.phien_id IN ({placeholders})
            WHERE sv.lop_id = %s AND sv.trang_thai = 1
            ORDER BY sv.mssv ASC
        """
        records = execute_query(sql, tuple(phien_ids) + (lop_id,))
    else:
        # Fallback: lấy tất cả SV, không có data điểm danh
        records = execute_query(
            """SELECT mssv, ho_ten, NULL as thoi_gian, NULL as status,
                      0 as late_minutes, NULL as gio_vao_lop, NULL as do_chinh_xac
               FROM sinh_vien WHERE lop_id = %s AND trang_thai = 1 ORDER BY mssv ASC""",
            (lop_id,),
        )

    wb = Workbook()

    # === Sheet 1: Chi tiết điểm danh ===
    ws1 = wb.active
    ws1.title = "Điểm Danh"

    # Styles
    header_fill = PatternFill(
        start_color="6A3CBC", end_color="6A3CBC", fill_type="solid"
    )
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=16, bold=True, color="333333")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header hệ thống
    ws1.merge_cells("A1:E1")
    ws1["A1"] = "HỆ THỐNG ĐIỂM DANH KHUÔN MẶT - MTUFACE"
    ws1["A1"].font = Font(name="Arial", size=11, bold=True, color="6A3CBC")
    ws1["A1"].alignment = Alignment(horizontal="left")

    # Tiêu đề
    ws1.merge_cells("A3:E3")
    ws1["A3"] = f"BẢNG ĐIỂM DANH - {lop.get('ten_lop', '')}".upper()
    ws1["A3"].font = title_font
    ws1["A3"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A4:E4")
    ws1["A4"] = (
        f"Mã lớp: {lop.get('ma_lop', '')} | Ngày: {date} | GV: {lop.get('giao_vien', '')}"
    )
    ws1["A4"].font = Font(name="Arial", size=12, italic=True)
    ws1["A4"].alignment = Alignment(horizontal="center")

    # Header hàng 6
    headers = [
        "STT",
        "MSSV",
        "Họ và Tên",
        "Giờ vào lớp",
        "Giờ điểm danh",
        "Đi trễ",
        "Trạng thái",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws1.row_dimensions[6].height = 25

    # Formatting fonts
    absent_font = Font(name="Arial", size=11, color="FF0000", bold=True)
    present_font = Font(name="Arial", size=11, color="008000")
    late_font = Font(name="Arial", size=11, color="E67E22", bold=True)
    excused_font = Font(name="Arial", size=11, color="2563EB", bold=True)

    total_present = 0
    total_late = 0
    total_absent = 0
    total_excused = 0

    for i, record in enumerate(records, 1):
        row = i + 6
        ws1.cell(row=row, column=1, value=i).border = thin_border
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws1.cell(row=row, column=2, value=record.get("mssv", "")).border = thin_border
        ws1.cell(row=row, column=3, value=record.get("ho_ten", "")).border = thin_border

        # Giờ vào lớp
        gio_vao_lop_td = record.get("gio_vao_lop")
        gio_vao_str = _format_time_value(gio_vao_lop_td) or "07:00"
        ws1.cell(row=row, column=4, value=gio_vao_str).border = thin_border
        ws1.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        # Dùng status và late_minutes từ DB (server đã tính)
        status = record.get("status")
        late_minutes = record.get("late_minutes", 0) or 0
        thoi_gian = record.get("thoi_gian")

        if thoi_gian:
            ws1.cell(
                row=row,
                column=5,
                value=(
                    thoi_gian.strftime("%H:%M:%S")
                    if hasattr(thoi_gian, "strftime")
                    else str(thoi_gian)
                ),
            ).border = thin_border
            ws1.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        else:
            ws1.cell(row=row, column=5, value="--:--:--").border = thin_border
            ws1.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        if status == AttendanceStatus.PRESENT:
            ws1.cell(row=row, column=6, value="Đúng giờ").border = thin_border
            status_cell = ws1.cell(row=row, column=7, value="Có mặt")
            status_cell.font = present_font
            total_present += 1
        elif status == AttendanceStatus.LATE:
            ws1.cell(row=row, column=6, value=f"{late_minutes} phút").border = (
                thin_border
            )
            ws1.cell(row=row, column=6).font = late_font
            status_cell = ws1.cell(row=row, column=7, value=f"Đi trễ ({late_minutes}p)")
            status_cell.font = late_font
            total_late += 1
        elif status == AttendanceStatus.EXCUSED_ABSENCE:
            ws1.cell(row=row, column=6, value="--").border = thin_border
            status_cell = ws1.cell(row=row, column=7, value="Vắng có phép")
            status_cell.font = excused_font
            total_excused += 1
        else:
            ws1.cell(row=row, column=6, value="--").border = thin_border
            status_cell = ws1.cell(row=row, column=7, value="Vắng mặt")
            status_cell.font = absent_font
            total_absent += 1

        ws1.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")

    # Auto-width
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 25
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 15
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 15

    # === Sheet 2: Thống kê ===
    ws2 = wb.create_sheet("Thống Kê")
    ws2["A1"] = "THỐNG KÊ ĐIỂM DANH"
    ws2["A1"].font = title_font

    si_so = len(records)
    ty_le = round((total_present + total_late) / max(1, si_so) * 100, 1)

    stats_data = [
        ("Sĩ số lớp", si_so),
        ("Có mặt", total_present),
        ("Đi trễ", total_late),
        ("Vắng có phép", total_excused),
        ("Vắng không phép", total_absent),
        ("Tỷ lệ chuyên cần", f"{ty_le}%"),
    ]
    for i, (label, value) in enumerate(stats_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)

    try:
        chart_buf = _create_attendance_pie_chart(
            total_present, total_late, total_absent
        )
        img = ExcelImage(chart_buf)
        ws2.add_image(img, "E3")
    except Exception as e:
        print(f"[EXCEL CHART ERROR] {e}")

    # Xuất ra bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _format_time_value(td_value):
    """Chuyển đổi timedelta/time/str thành chuỗi HH:MM."""
    if td_value is None:
        return None
    if isinstance(td_value, str):
        return td_value[:5]
    if hasattr(td_value, "total_seconds"):
        total_seconds = int(td_value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    if hasattr(td_value, "strftime"):
        return td_value.strftime("%H:%M")
    return str(td_value)[:5]


def to_pdf(lop_id, date=None):
    """
    Tạo file PDF điểm danh.
    Query theo phiên (phien_id) thay vì theo ngày.

    Args:
        lop_id: ID lớp học
        date: Ngày xuất

    Returns:
        bytes: Nội dung file PDF
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        Image as RLImage,
    )
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from services.attendance_policy import AttendanceStatus

    try:
        pdfmetrics.registerFont(TTFont("Arial", "arial.ttf"))
        pdfmetrics.registerFont(TTFont("Arial-Bold", "arialbd.ttf"))
        font_regular = "Arial"
        font_bold = "Arial-Bold"
    except Exception as e:
        print("[PDF FONT ERROR]", e)
        font_regular = "Helvetica"
        font_bold = "Helvetica-Bold"

    lop = class_service.get_by_id(lop_id)
    if not lop:
        return None

    if not date:
        date = datetime.now().strftime("%Y-%m-%d")

    # Tìm phiên đã chốt trong ngày cho lớp này
    session_row = execute_query(
        """SELECT id FROM phien_diem_danh
           WHERE lop_id = %s AND CAST(bat_dau AS DATE) = %s
           AND ISNULL(is_cancelled, 0) = 0
           ORDER BY bat_dau DESC""",
        (lop_id, date),
    )

    if session_row:
        phien_ids = [s["id"] for s in session_row]
        placeholders = ",".join(["%s"] * len(phien_ids))
        sql = f"""
            SELECT sv.mssv, sv.ho_ten,
                   dd.thoi_gian, dd.status, dd.late_minutes, dd.gio_vao_lop
            FROM sinh_vien sv
            LEFT JOIN diem_danh dd ON sv.id = dd.sinh_vien_id
                AND dd.phien_id IN ({placeholders})
            WHERE sv.lop_id = %s AND sv.trang_thai = 1
            ORDER BY sv.mssv ASC
        """
        records = execute_query(sql, tuple(phien_ids) + (lop_id,))
    else:
        records = execute_query(
            """SELECT mssv, ho_ten, NULL as thoi_gian, NULL as status,
                      0 as late_minutes, NULL as gio_vao_lop
               FROM sinh_vien WHERE lop_id = %s AND trang_thai = 1 ORDER BY mssv ASC""",
            (lop_id,),
        )

    # Tạo PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()

    styles["Normal"].fontName = font_regular
    styles["Title"].fontName = font_bold
    styles["Heading1"].fontName = font_bold
    styles["Heading2"].fontName = font_bold

    elements = []

    elements.append(
        Paragraph("<b>HE THONG DIEM DANH KHUON MAT - MTUFACE</b>", styles["Normal"])
    )
    elements.append(Spacer(1, 0.5 * cm))

    title_style = styles["Title"]
    title_style.textColor = colors.HexColor("#6A3CBC")
    elements.append(Paragraph(f"<b>BANG DIEM DANH LOP HOC</b>", title_style))
    elements.append(
        Paragraph(
            f"<b><font size='14'>{lop.get('ten_lop', '')}</font></b>", styles["Title"]
        )
    )

    elements.append(Spacer(1, 0.2 * cm))
    elements.append(
        Paragraph(
            f"<b>Ma lop:</b> {lop.get('ma_lop', '')} &nbsp;&nbsp;&nbsp;&nbsp; <b>Ngay:</b> {date} &nbsp;&nbsp;&nbsp;&nbsp; <b>Giang vien:</b> {lop.get('giao_vien', '')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 0.8 * cm))

    # Bảng dữ liệu — dùng status enum từ DB
    table_data = [["STT", "MSSV", "Ho va Ten", "Vao lop", "Diem danh", "Trang thai"]]
    total_present = 0
    total_late = 0
    total_absent = 0

    for i, record in enumerate(records, 1):
        thoi_gian = record.get("thoi_gian")
        status = record.get("status")
        late_minutes = record.get("late_minutes", 0) or 0
        gio_vao_str = _format_time_value(record.get("gio_vao_lop")) or "07:00"

        if thoi_gian:
            tg_str = (
                thoi_gian.strftime("%H:%M:%S")
                if hasattr(thoi_gian, "strftime")
                else str(thoi_gian)
            )
        else:
            tg_str = "--:--:--"

        if status == AttendanceStatus.PRESENT:
            trang_thai_str = "Co mat"
            total_present += 1
        elif status == AttendanceStatus.LATE:
            trang_thai_str = f"Tre {late_minutes} p"
            total_late += 1
        elif status == AttendanceStatus.EXCUSED_ABSENCE:
            trang_thai_str = "Vang co phep"
            total_absent += 1
        else:
            trang_thai_str = "Vang mat"
            total_absent += 1

        table_data.append(
            [
                str(i),
                record.get("mssv", ""),
                record.get("ho_ten", ""),
                gio_vao_str,
                tg_str,
                trang_thai_str,
            ]
        )

    try:
        chart_buf = _create_attendance_pie_chart(
            total_present, total_late, total_absent
        )
        chart_img = RLImage(chart_buf, width=11 * cm, height=6.5 * cm)
        elements.append(chart_img)
        elements.append(Spacer(1, 0.5 * cm))
    except Exception as e:
        print(f"[PDF CHART ERROR] {e}")

    if len(table_data) > 1:
        table = Table(
            table_data,
            colWidths=[1.2 * cm, 3.2 * cm, 4.5 * cm, 2.3 * cm, 3 * cm, 3.5 * cm],
        )

        style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6A3CBC")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTNAME", (0, 1), (-1, -1), font_regular),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#F8FAFC")],
                ),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )

        for row_idx, row_data in enumerate(table_data[1:], start=1):
            st_val = row_data[5]
            if st_val == "Co mat":
                style.add(
                    "TEXTCOLOR", (5, row_idx), (5, row_idx), colors.HexColor("#16A34A")
                )
            elif st_val == "Vang mat" or st_val == "Vang co phep":
                style.add(
                    "TEXTCOLOR", (5, row_idx), (5, row_idx), colors.HexColor("#DC2626")
                )
            else:
                style.add(
                    "TEXTCOLOR", (5, row_idx), (5, row_idx), colors.HexColor("#D97706")
                )

        table.setStyle(style)
        elements.append(table)
    else:
        elements.append(Paragraph("Khong co du lieu diem danh.", styles["Normal"]))

    # Chữ ký
    elements.append(Spacer(1, 2 * cm))
    sig_data = [
        ["", f"Ngay {date}"],
        ["Giao vien", "Nguoi lap bieu"],
        ["", ""],
        [lop.get("giao_vien", ""), "Admin"],
    ]
    sig_table = Table(sig_data, colWidths=[8 * cm, 8 * cm])
    sig_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, -1), font_regular),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
            ]
        )
    )
    elements.append(sig_table)

    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def roster_to_excel(lop_id):
    """
    Xuất danh sách lớp trắng (Roster) để Giảng viên sử dụng ngoài hệ thống.
    Chỉ bao gồm: STT, MSSV, Họ Tên, Giới Tính, Ngày Sinh, Ghi Chú (trống).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    lop = class_service.get_by_id(lop_id)
    if not lop:
        return None

    sql = """
        SELECT mssv, ho_ten, gioi_tinh, ngay_sinh
        FROM sinh_vien
        WHERE lop_id = %s AND trang_thai = 1
        ORDER BY mssv ASC
    """
    students = execute_query(sql, (lop_id,))

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh Sách Lớp"

    header_fill = PatternFill(
        start_color="6A3CBC", end_color="6A3CBC", fill_type="solid"
    )
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=16, bold=True, color="333333")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header hệ thống
    ws.merge_cells("A1:F1")
    ws["A1"] = "HỆ THỐNG ĐIỂM DANH KHUÔN MẶT - MTUFACE"
    ws["A1"].font = Font(name="Arial", size=11, bold=True, color="6A3CBC")
    ws["A1"].alignment = Alignment(horizontal="left")

    # Tiêu đề
    ws.merge_cells("A3:F3")
    ws["A3"] = f"DANH SÁCH SINH VIÊN - {lop.get('ten_lop', '')}".upper()
    ws["A3"].font = title_font
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:F4")
    ws["A4"] = (
        f"Mã lớp: {lop.get('ma_lop', '')} | GV: {lop.get('giao_vien', '')} | Sĩ số: {len(students)}"
    )
    ws["A4"].font = Font(name="Arial", size=11, italic=True)
    ws["A4"].alignment = Alignment(horizontal="center")

    headers = ["STT", "MSSV", "Họ và Tên", "Giới Tính", "Ngày Sinh", "Ghi Chú"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[6].height = 25

    for i, sv in enumerate(students, 1):
        row = i + 6
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=sv.get("mssv", "")).border = thin_border
        ws.cell(row=row, column=3, value=sv.get("ho_ten", "")).border = thin_border

        gioi_tinh = sv.get("gioi_tinh")
        gt_str = "Nam" if gioi_tinh == 1 else ("Nữ" if gioi_tinh == 0 else "")
        ws.cell(row=row, column=4, value=gt_str).border = thin_border

        ngay_sinh = sv.get("ngay_sinh")
        ns_str = ngay_sinh.strftime("%d/%m/%Y") if ngay_sinh else ""
        ws.cell(row=row, column=5, value=ns_str).border = thin_border

        ws.cell(row=row, column=6, value="").border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def monthly_matrix_to_excel(lop_id, month, year):
    """
    Xuất ma trận điểm danh theo tháng (cập nhật sử dụng export_semester_report).
    """
    return export_semester_report(lop_id, month=month, year=year)


def export_session_report(phien_id):
    """
    Xuất báo cáo chi tiết cho một phiên điểm danh chuẩn xác theo phien_id (.xlsx),
    hiển thị đầy đủ si_so_chot, trạng thái P/L/E/A, tỷ lệ tham dự & điểm chuyên cần.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from db.connection import execute_one, execute_query

    phien = execute_one(
        "SELECT p.*, l.ma_lop, l.ten_lop, l.giao_vien FROM phien_diem_danh p JOIN lop_hoc l ON p.lop_id = l.id WHERE p.id = %s",
        (phien_id,),
    )
    if not phien:
        return None

    lop_id = phien["lop_id"]
    sql = """
        SELECT sv.id, sv.mssv, sv.ho_ten,
               d.thoi_gian, d.trang_thai, d.status, d.late_minutes, d.ghi_chu
        FROM sinh_vien sv
        LEFT JOIN diem_danh d ON sv.id = d.sinh_vien_id AND d.phien_id = %s
        WHERE sv.lop_id = %s AND sv.trang_thai = 1
        ORDER BY sv.mssv ASC
    """
    records = execute_query(sql, (phien_id, lop_id))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Phiên {phien_id}"

    header_fill = PatternFill(
        start_color="6A3CBC", end_color="6A3CBC", fill_type="solid"
    )
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=15, bold=True, color="333333")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Đếm trạng thái
    p_count = sum(1 for r in records if r.get("status") == "PRESENT")
    l_count = sum(1 for r in records if r.get("status") == "LATE")
    e_count = sum(1 for r in records if r.get("status") == "EXCUSED_ABSENCE")
    a_count = sum(
        1
        for r in records
        if r.get("status") == "UNEXCUSED_ABSENCE" or not r.get("status")
    )

    total_sv = len(records)
    si_so_chot = phien.get("si_so_chot") or total_sv
    ty_le_tham_du = round((p_count + l_count) / max(1, si_so_chot) * 100, 1)
    diem_chuyen_can = round(
        (p_count * 1.0 + l_count * 0.8 + e_count * 0.5) / max(1, si_so_chot) * 100, 1
    )

    # Tiêu đề báo cáo
    ws.merge_cells("A1:H1")
    ws["A1"] = "HỆ THỐNG ĐIỂM DANH KHUÔN MẶT - MTUFACE"
    ws["A1"].font = Font(name="Arial", size=11, bold=True, color="6A3CBC")

    ws.merge_cells("A3:H3")
    ws["A3"] = f"BÁO CÁO PHIÊN ĐIỂM DANH - {phien.get('ten_lop', '')}".upper()
    ws["A3"].font = title_font
    ws["A3"].alignment = Alignment(horizontal="center")

    bat_dau_str = (
        phien["bat_dau"].strftime("%d/%m/%Y %H:%M") if phien.get("bat_dau") else "N/A"
    )
    ws.merge_cells("A4:H4")
    ws["A4"] = (
        f"Mã lớp: {phien.get('ma_lop', '')} | Bắt đầu: {bat_dau_str} | Sĩ số chốt: {si_so_chot} | Tỷ lệ tham dự: {ty_le_tham_du}% | Điểm chuyên cần: {diem_chuyen_can}%"
    )
    ws["A4"].font = Font(name="Arial", size=11, italic=True)
    ws["A4"].alignment = Alignment(horizontal="center")

    headers = [
        "STT",
        "MSSV",
        "Họ và Tên",
        "Giờ điểm danh",
        "Đi trễ (phút)",
        "Ký hiệu",
        "Trạng thái",
        "Ghi chú",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[6].height = 25

    font_p = Font(name="Arial", size=11, color="16A34A", bold=True)
    font_l = Font(name="Arial", size=11, color="D97706", bold=True)
    font_e = Font(name="Arial", size=11, color="2563EB", bold=True)
    font_a = Font(name="Arial", size=11, color="DC2626", bold=True)

    for idx, r in enumerate(records, 1):
        row = idx + 6
        status = r.get("status", "UNEXCUSED_ABSENCE")
        late_min = r.get("late_minutes", 0)

        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=r.get("mssv", "")).border = thin_border
        ws.cell(row=row, column=3, value=r.get("ho_ten", "")).border = thin_border

        tg_str = (
            r["thoi_gian"].strftime("%H:%M:%S") if r.get("thoi_gian") else "--:--:--"
        )
        ws.cell(row=row, column=4, value=tg_str).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        late_str = f"{late_min}p" if late_min > 0 else "--"
        ws.cell(row=row, column=5, value=late_str).border = thin_border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        if status == "PRESENT":
            symbol, label, font_style = "P", "Có mặt đúng giờ", font_p
        elif status == "LATE":
            symbol, label, font_style = (
                f"L({late_min})",
                f"Đi trễ {late_min} phút",
                font_l,
            )
        elif status == "EXCUSED_ABSENCE":
            symbol, label, font_style = "E", "Vắng có phép", font_e
        else:
            symbol, label, font_style = "A", "Vắng không phép", font_a

        c_sym = ws.cell(row=row, column=6, value=symbol)
        c_sym.border = thin_border
        c_sym.font = font_style
        c_sym.alignment = Alignment(horizontal="center")

        c_lbl = ws.cell(row=row, column=7, value=label)
        c_lbl.border = thin_border
        c_lbl.font = font_style

        ws.cell(row=row, column=8, value=r.get("ghi_chu", "")).border = thin_border

    # Bảng chú giải Legend ở dưới
    leg_row = len(records) + 9
    ws.cell(row=leg_row, column=2, value="CHÚ GIẢI KÝ HIỆU (LEGEND):").font = Font(
        name="Arial", size=11, bold=True
    )
    legends = [
        ("P", "Có mặt đúng giờ (Present)", font_p),
        ("L(x)", "Đi trễ x phút (Late)", font_l),
        ("E", "Vắng có phép (Excused Absence)", font_e),
        ("A", "Vắng không phép (Unexcused Absence)", font_a),
        (
            "R",
            "Đơn xin phép bị từ chối (Rejected)",
            Font(name="Arial", size=11, color="7F1D1D"),
        ),
        (
            "?",
            "Đơn xin phép đang chờ duyệt (Pending)",
            Font(name="Arial", size=11, color="4B5563"),
        ),
        (
            "C",
            "Buổi học / phiên bị hủy (Cancelled)",
            Font(name="Arial", size=11, color="9CA3AF"),
        ),
    ]
    for i, (sym, desc, f_style) in enumerate(legends):
        r_l = leg_row + i + 1
        c_s = ws.cell(row=r_l, column=2, value=sym)
        c_s.font = f_style
        c_s.alignment = Alignment(horizontal="center")
        ws.cell(row=r_l, column=3, value=desc).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_semester_report(lop_id, month=None, year=None):
    """
    Xuất báo cáo ma trận điểm danh các phiên đã đóng theo lớp (.xlsx),
    chuẩn hóa ký hiệu P, L(x), E, A, R, ?, C kèm bảng chú giải Legend.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from db.connection import execute_one, execute_query
    from openpyxl.utils import get_column_letter

    lop = class_service.get_by_id(lop_id)
    if not lop:
        return None

    sql_sess = (
        "SELECT id, bat_dau FROM phien_diem_danh WHERE lop_id = %s AND trang_thai = 0"
    )
    params_sess = [lop_id]
    if month and year:
        sql_sess += " AND MONTH(bat_dau) = %s AND YEAR(bat_dau) = %s"
        params_sess.extend([month, year])
    sql_sess += " ORDER BY bat_dau ASC"
    sessions = execute_query(sql_sess, params_sess)

    students = execute_query(
        "SELECT id, mssv, ho_ten FROM sinh_vien WHERE lop_id = %s AND trang_thai = 1 ORDER BY mssv ASC",
        (lop_id,),
    )
    if not students:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"Chuyên Cần Lớp {lop.get('ma_lop', '')}"

    header_fill = PatternFill(
        start_color="6A3CBC", end_color="6A3CBC", fill_type="solid"
    )
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="333333")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    num_sessions = len(sessions)
    last_col = (
        3 + num_sessions + 6
    )  # STT + MSSV + HoTen + sessions + P + L + E + A + TyLe + DiemCC

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = "HỆ THỐNG ĐIỂM DANH KHUÔN MẶT - MTUFACE"
    ws["A1"].font = Font(name="Arial", size=11, bold=True, color="6A3CBC")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    title_str = f"BẢNG CHUYÊN CẦN LỚP {lop.get('ten_lop', '')}".upper()
    if month and year:
        title_str += f" - THÁNG {month}/{year}"
    ws["A3"] = title_str
    ws["A3"].font = title_font
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    ws["A4"] = (
        f"Mã lớp: {lop.get('ma_lop', '')} | GV: {lop.get('giao_vien', '')} | Tổng phiên: {num_sessions}"
    )
    ws["A4"].font = Font(name="Arial", size=11, italic=True)
    ws["A4"].alignment = Alignment(horizontal="center")

    headers = ["STT", "MSSV", "Họ và Tên"]
    for idx, sess in enumerate(sessions, 1):
        d_str = sess["bat_dau"].strftime("%d/%m") if sess.get("bat_dau") else f"P{idx}"
        headers.append(f"P{idx} ({d_str})")
    headers.extend(["P", "L", "E", "A", "Tỷ lệ TD %", "Điểm CC %"])

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[6].height = 25

    # Map dữ liệu điểm danh (sinh_vien_id, phien_id) -> row
    dd_records = execute_query(
        "SELECT sinh_vien_id, phien_id, status, late_minutes FROM diem_danh WHERE lop_id = %s",
        (lop_id,),
    )
    dd_map = {(r["sinh_vien_id"], r["phien_id"]): r for r in dd_records}

    font_p = Font(name="Arial", size=10, color="16A34A", bold=True)
    font_l = Font(name="Arial", size=10, color="D97706", bold=True)
    font_e = Font(name="Arial", size=10, color="2563EB", bold=True)
    font_a = Font(name="Arial", size=10, color="DC2626", bold=True)

    for i, sv in enumerate(students, 1):
        row = i + 6
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=sv.get("mssv", "")).border = thin_border
        ws.cell(row=row, column=3, value=sv.get("ho_ten", "")).border = thin_border

        p_cnt = l_cnt = e_cnt = a_cnt = 0
        for s_idx, sess in enumerate(sessions, 1):
            col = 3 + s_idx
            c_cell = ws.cell(row=row, column=col)
            c_cell.border = thin_border
            c_cell.alignment = Alignment(horizontal="center")

            key = (sv["id"], sess["id"])
            if key in dd_map:
                st = dd_map[key].get("status")
                late_m = dd_map[key].get("late_minutes", 0)
                if st == "PRESENT":
                    c_cell.value = "P"
                    c_cell.font = font_p
                    p_cnt += 1
                elif st == "LATE":
                    c_cell.value = f"L({late_m})"
                    c_cell.font = font_l
                    l_cnt += 1
                elif st == "EXCUSED_ABSENCE":
                    c_cell.value = "E"
                    c_cell.font = font_e
                    e_cnt += 1
                else:
                    c_cell.value = "A"
                    c_cell.font = font_a
                    a_cnt += 1
            else:
                c_cell.value = "A"
                c_cell.font = font_a
                a_cnt += 1

        # Tổng P, L, E, A, TyLe, DiemCC
        base_col = 3 + num_sessions
        ws.cell(row=row, column=base_col + 1, value=p_cnt).border = thin_border
        ws.cell(row=row, column=base_col + 2, value=l_cnt).border = thin_border
        ws.cell(row=row, column=base_col + 3, value=e_cnt).border = thin_border
        ws.cell(row=row, column=base_col + 4, value=a_cnt).border = thin_border

        ty_le_td = (
            round((p_cnt + l_cnt) / max(1, num_sessions) * 100, 1)
            if num_sessions > 0
            else 0
        )
        diem_cc = (
            round(
                (p_cnt * 1.0 + l_cnt * 0.8 + e_cnt * 0.5) / max(1, num_sessions) * 100,
                1,
            )
            if num_sessions > 0
            else 0
        )

        c_td = ws.cell(row=row, column=base_col + 5, value=f"{ty_le_td}%")
        c_td.border = thin_border
        c_td.alignment = Alignment(horizontal="center")

        c_cc = ws.cell(row=row, column=base_col + 6, value=f"{diem_cc}%")
        c_cc.border = thin_border
        c_cc.alignment = Alignment(horizontal="center")
        c_cc.font = Font(name="Arial", size=10, bold=True, color="6A3CBC")

    # Bảng chú giải Legend
    leg_row = len(students) + 9
    ws.cell(row=leg_row, column=2, value="CHÚ GIẢI KÝ HIỆU (LEGEND):").font = Font(
        name="Arial", size=11, bold=True
    )
    legends = [
        ("P", "Có mặt đúng giờ (Present)", font_p),
        ("L(x)", "Đi trễ x phút (Late)", font_l),
        ("E", "Vắng có phép (Excused Absence)", font_e),
        ("A", "Vắng không phép (Unexcused Absence)", font_a),
        (
            "R",
            "Đơn xin phép bị từ chối (Rejected)",
            Font(name="Arial", size=10, color="7F1D1D"),
        ),
        (
            "?",
            "Đơn xin phép đang chờ duyệt (Pending)",
            Font(name="Arial", size=10, color="4B5563"),
        ),
        (
            "C",
            "Buổi học / phiên bị hủy (Cancelled)",
            Font(name="Arial", size=10, color="9CA3AF"),
        ),
    ]
    for i, (sym, desc, f_style) in enumerate(legends):
        r_l = leg_row + i + 1
        c_s = ws.cell(row=r_l, column=2, value=sym)
        c_s.font = f_style
        c_s.alignment = Alignment(horizontal="center")
        ws.cell(row=r_l, column=3, value=desc).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    for idx in range(1, num_sessions + 1):
        ws.column_dimensions[get_column_letter(3 + idx)].width = 10
    ws.column_dimensions[get_column_letter(3 + num_sessions + 1)].width = 7
    ws.column_dimensions[get_column_letter(3 + num_sessions + 2)].width = 7
    ws.column_dimensions[get_column_letter(3 + num_sessions + 3)].width = 7
    ws.column_dimensions[get_column_letter(3 + num_sessions + 4)].width = 7
    ws.column_dimensions[get_column_letter(3 + num_sessions + 5)].width = 13
    ws.column_dimensions[get_column_letter(3 + num_sessions + 6)].width = 13

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_filename(lop_id, date, ext):
    """Tạo tên file xuất dữ liệu."""
    lop = class_service.get_by_id(lop_id)
    ma_lop = lop.get("ma_lop", "UNKNOWN") if lop else "UNKNOWN"
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    return f"DiemDanh_{ma_lop}_{date}.{ext}"
